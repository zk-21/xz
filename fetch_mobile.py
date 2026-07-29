#!/usr/bin/env python3
"""
大乐透数据下载工具 - 手机版
不需要电脑！在手机上用 Pydroid 3 / Termux 直接运行即可。

用法：python fetch_mobile.py [期数]
  例：python fetch_mobile.py        # 默认120期
  例：python fetch_mobile.py 500    # 下载500期

输出文件保存在脚本同目录下：
  - all_draws.js    JavaScript格式
  - 1.xlsx          Excel报表
"""
import urllib.request
import urllib.error
import json
import os
import sys
import time

API_BASE = "https://webapi.sporttery.cn/gateway/lottery/"
PAGE_SIZE = 30
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer": "https://www.lottery.gov.cn/",
}


def fetch_page(page_no, page_size=PAGE_SIZE):
    """抓取一页开奖数据"""
    params = f"getHistoryPageListV1.qry?gameNo=85&provinceId=0&pageSize={page_size}&isVerify=1&pageNo={page_no}"
    url = API_BASE + params
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        if not data.get("success"):
            raise Exception(data.get("errorMessage", "API返回错误"))
        return data["value"]["list"]
    except urllib.error.HTTPError as e:
        raise Exception(f"HTTP错误 {e.code}")
    except urllib.error.URLError as e:
        raise Exception(f"网络错误: {e.reason}\n请检查手机网络连接")


def fetch_all(period):
    """抓取指定期数的数据"""
    total_pages = (period + PAGE_SIZE - 1) // PAGE_SIZE
    seen = set()
    draws = []

    print(f"目标: {period} 期, 共 {total_pages} 页\n")

    for p in range(1, total_pages + 1):
        pct = (p - 1) * 100 // total_pages
        bar = "#" * (pct // 5) + "-" * (20 - pct // 5)
        print(f"\r[{bar}] {pct}%  第 {p}/{total_pages} 页", end="", flush=True)

        items = fetch_page(p)
        if not items:
            break

        for item in items:
            issue = item["lotteryDrawNum"]
            if issue in seen:
                continue
            seen.add(issue)

            nums = [int(x) for x in item["lotteryDrawResult"].split()]
            if len(nums) < 7:
                continue

            draws.append({
                "issue": issue,
                "date": item.get("lotteryDrawTime", ""),
                "front": sorted(nums[:5]),
                "back": sorted(nums[5:7]),
            })

        if len(draws) >= period:
            break
        time.sleep(0.2)

    print(f"\r[####################] 100%  完成!       ")

    # 降序排列（最新在前）
    draws.sort(key=lambda d: int(d["issue"]), reverse=True)
    return draws[:period]


def save_js(draws, filepath):
    """生成 all_draws.js"""
    lines = [
        f"// 大乐透开奖数据 — {len(draws)}期",
        f"// 生成时间: {time.strftime('%Y-%m-%d %H:%M:%S')}",
        "// 数据来源: 中国体彩官方API",
        "",
        "const ALL_DRAWS = [",
    ]

    # 升序写入（最旧在前）
    for i, d in enumerate(reversed(draws)):
        front_str = " ".join(f"{n:02d}" for n in d["front"])
        back_str = " ".join(f"{n:02d}" for n in d["back"])
        line = (
            f'  {{issue:"{d["issue"]}",date:"{d["date"]}",'
            f'front:[{",".join(map(str, d["front"]))}],'
            f'back:[{",".join(map(str, d["back"]))}],'
            f'front_str:"{front_str}",back_str:"{back_str}"}}'
        )
        if i < len(draws) - 1:
            line += ","
        lines.append(line)

    lines.append("];\n")

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"\n  all_draws.js 已生成 ({os.path.getsize(filepath):,} bytes)")


def save_xlsx(draws, filepath):
    """生成 1.xlsx"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("\n  ⚠ openpyxl 未安装，尝试安装中...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "大乐透开奖数据"

    # 表头
    headers = [
        "期号", "日期", "星期",
        "前1", "前2", "前3", "前4", "前5", "后1", "后2",
        "一等注数", "一等金额", "一等追加注数", "一等追加金额",
        "二等注数", "二等金额", "二等追加注数", "二等追加金额",
        "三等注数", "三等金额", "四等注数", "四等金额",
        "五等注数", "五等金额", "六等注数", "六等金额",
        "销售额", "奖池",
    ]
    header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    front_font = Font(name="微软雅黑", size=11, bold=True, color="C0392B")
    back_font = Font(name="微软雅黑", size=11, bold=True, color="2980B9")
    normal_font = Font(name="微软雅黑", size=10)
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    # 数据行
    weekdays = ["日", "一", "二", "三", "四", "五", "六"]
    for r, d in enumerate(draws, 2):
        ws.cell(row=r, column=1, value=d["issue"]).font = normal_font
        ws.cell(row=r, column=2, value=d["date"]).font = normal_font

        # 星期
        try:
            t = time.strptime(d["date"], "%Y-%m-%d")
            wd = weekdays[t.tm_wday]
        except:
            wd = ""
        ws.cell(row=r, column=3, value=wd).font = normal_font

        # 前区（红色加粗）
        for i, n in enumerate(d["front"]):
            ws.cell(row=r, column=4 + i, value=n).font = front_font

        # 后区（蓝色加粗）
        for i, n in enumerate(d["back"]):
            ws.cell(row=r, column=10 + i, value=n).font = back_font

        # 居中
        for c in range(1, 29):
            ws.cell(row=r, column=c).alignment = center
            ws.cell(row=r, column=c).border = thin_border

    # 列宽
    widths = [8, 12, 5, 6, 6, 6, 6, 6, 6, 6,
              10, 10, 10, 10, 10, 10, 10, 10,
              10, 10, 10, 10, 10, 10, 10, 10, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 冻结首行 + 自动筛选
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(draws) + 1}"

    wb.save(filepath)
    print(f"  1.xlsx 已生成 ({os.path.getsize(filepath):,} bytes)")


def main():
    period = 120
    if len(sys.argv) > 1:
        period = int(sys.argv[1])
    period = min(period, 2000)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    os.chdir(script_dir)

    print("=" * 44)
    print("  大乐透数据下载工具（手机版）")
    print("=" * 44)
    print(f"  下载期数: {period} 期")
    print(f"  保存位置: {script_dir}")
    print()

    try:
        draws = fetch_all(period)
        print(f"\n  获取 {len(draws)} 期数据")
        print(f"  范围: {draws[-1]['issue']} ~ {draws[0]['issue']}")
        print()

        save_js(draws, os.path.join(script_dir, "all_draws.js"))
        save_xlsx(draws, os.path.join(script_dir, "1.xlsx"))

        print(f"\n{'=' * 44}")
        print("  下载完成！文件保存在：")
        print(f"  {script_dir}")
        print(f"{'=' * 44}")

    except Exception as e:
        print(f"\n  错误: {e}")
        print("\n  常见问题：")
        print("  1. 检查手机网络是否正常")
        print("  2. 检查是否被防火墙拦截")
        print("  3. 尝试切换 WiFi / 4G")
        sys.exit(1)


if __name__ == "__main__":
    main()
