"""
GitHub Actions / 本地 通用数据更新脚本
策略：多数据源自动切换
  1. 中国体彩 API (webapi.sporttery.cn) - 主要来源 (连续无间断)
  2. 500彩票网 (datachart.500.com) - 备用 (API被阻时使用)
  3. 官网网页爬虫 (www.lottery.gov.cn) - 兜底
"""
import json
import os
import re
import sys
import time
import ssl
import urllib.request
import urllib.error
import urllib.parse
import warnings

warnings.filterwarnings("ignore", message="Unverified HTTPS request")
try:
    import urllib3
    warnings.filterwarnings("ignore", category=urllib3.exceptions.InsecureRequestWarning)
except ImportError:
    pass

PROXY = os.environ.get("HTTP_PROXY") or os.environ.get("http_proxy") or os.environ.get("HTTPS_PROXY") or os.environ.get("https_proxy")
if PROXY:
    print(f"  使用代理: {PROXY}")

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

PERIOD = 1000
PAGE_SIZE = 30
TOTAL_PAGES = 34
OUT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
HTML_HEADERS = {
    "User-Agent": UA,
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
    "Connection": "keep-alive",
}
API_HEADERS = {
    "User-Agent": UA,
    "Referer": "https://www.lottery.gov.cn/",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
    "Connection": "keep-alive",
}


def _http_get(url, headers=None, timeout=60):
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    req = urllib.request.Request(url, headers=headers or HTML_HEADERS)
    with urllib.request.urlopen(req, timeout=timeout, context=ctx) as resp:
        return resp.read().decode("utf-8", errors="replace")


def _requests_get(url, params=None, headers=None, timeout=60):
    kwargs = {"headers": headers or HTML_HEADERS, "timeout": timeout, "verify": False}
    if PROXY:
        kwargs["proxies"] = {"http": PROXY, "https": PROXY}
    if params:
        kwargs["params"] = params
    resp = requests.get(url, **kwargs)
    return resp


# ============ 数据源 1: 中国体彩 API (主) ============

API_URLS = [
    "https://webapi.sporttery.cn/gateway/lottery/getHistoryPageListV1.qry",
    "http://webapi.sporttery.cn/gateway/lottery/getHistoryPageListV1.qry",
]


def fetch_from_api(timeout=60):
    """通过官方 API 获取全部数据 (连续无间断, 约1020期)"""
    draws = []
    seen = set()

    print("  尝试中国体彩 API...")
    for page_no in range(1, TOTAL_PAGES + 1):
        params = {
            "gameNo": "85",
            "provinceId": "0",
            "pageSize": str(PAGE_SIZE),
            "isVerify": "1",
            "pageNo": str(page_no),
        }

        got_data = False
        for url in API_URLS:
            try:
                if HAS_REQUESTS:
                    resp = _requests_get(url, params, API_HEADERS, timeout)
                    if resp.status_code == 567:
                        print(f"    API 被阻止 (567)")
                        return []
                    resp.raise_for_status()
                    data = resp.json()
                else:
                    query = urllib.parse.urlencode(params)
                    raw = _http_get(url + "?" + query, API_HEADERS, timeout)
                    data = json.loads(raw)

                if data.get("success"):
                    items = data.get("value", {}).get("list", [])
                    if not items:
                        got_data = True
                        break
                    for item in items:
                        issue = str(item.get("lotteryDrawNum", ""))
                        if issue in seen:
                            continue
                        seen.add(issue)
                        nums_raw = item.get("lotteryDrawResult", "")
                        if isinstance(nums_raw, str):
                            nums = [int(x) for x in nums_raw.split() if x.strip().isdigit()]
                        elif isinstance(nums_raw, list):
                            nums = [int(x) for x in nums_raw]
                        else:
                            nums = []
                        if issue and len(nums) >= 7:
                            draws.append({
                                "issue": issue,
                                "date": item.get("lotteryDrawTime", ""),
                                "front": sorted(nums[:5]),
                                "back": sorted(nums[5:7]),
                            })
                    got_data = True
                    break
                else:
                    continue
            except Exception as e:
                if "567" in str(e):
                    print(f"    API 被阻止 (567)")
                    return []
                continue

        if not got_data:
            break

        if page_no == 1:
            print(f"    第1页: 30 期")
        if page_no % 5 == 0:
            print(f"    第{page_no}页: 共 {len(draws)} 期")

        time.sleep(0.3)

    draws.sort(key=lambda d: int(d["issue"]), reverse=True)
    print(f"    API 共获取 {len(draws)} 期")
    return draws


# ============ 数据源 2: 500彩票网 ============

def fetch_from_500(start_period=19000, timeout=60):
    """从 500.com 抓取大乐透历史数据 (有年度间隙, 约971期)"""
    draws = []
    url = f"https://datachart.500.com/dlt/history/newinc/history.php?start={start_period:05d}&end=99999"

    print(f"  尝试 500彩票网: {url[:80]}...")
    try:
        raw = _http_get(url, HTML_HEADERS, timeout)
        print(f"    获取成功: {len(raw)} bytes")

        trs = re.findall(r'<tr[^>]*>(.*?)</tr>', raw, re.DOTALL)
        for tr in trs:
            td_detail = re.findall(r'<td[^>]*>(.*?)</td>', tr, re.DOTALL)
            if len(td_detail) < 16:
                continue
            clean = [re.sub(r'<[^>]+>', '', td).strip() for td in td_detail]

            issue = clean[1]
            if not re.match(r'^\d{5}$', issue):
                continue

            front_nums = []
            for j in range(2, 7):
                v = clean[j]
                if re.match(r'^\d{1,2}$', v):
                    n = int(v)
                    if 1 <= n <= 35:
                        front_nums.append(n)

            back_nums = []
            for j in range(7, 9):
                v = clean[j]
                if re.match(r'^\d{1,2}$', v):
                    n = int(v)
                    if 1 <= n <= 35:
                        back_nums.append(n)

            date_str = clean[15] if len(clean) > 15 else ""
            if not re.match(r'\d{4}-\d{2}-\d{2}', date_str):
                date_str = ""

            if len(front_nums) == 5 and len(back_nums) == 2 and date_str:
                draws.append({
                    "issue": issue,
                    "date": date_str,
                    "front": sorted(front_nums),
                    "back": sorted(back_nums),
                })

        if draws:
            draws.sort(key=lambda d: int(d["issue"]), reverse=True)
            print(f"    解析到 {len(draws)} 期 (最新 {draws[0]['issue']})")
        else:
            print(f"    未解析到数据")
    except Exception as e:
        print(f"    失败: {type(e).__name__}: {str(e)[:80]}")

    return draws


# ============ 数据源 3: 官网网页爬虫 ============

def fetch_from_webpage(timeout=60):
    """从体彩官网页面抓取（兜底方案）"""
    draws = []
    urls = [
        "https://www.lottery.gov.cn/kj/kjlb.html?dlt",
        "http://www.lottery.gov.cn/kj/kjlb.html?dlt",
    ]

    for url in urls:
        try:
            print(f"  尝试官网页面: {url}")
            raw = _http_get(url, HTML_HEADERS, timeout)

            for pattern in [
                r'var\s+kjData\s*=\s*(\[.*?\]);',
                r'var\s+lotteryData\s*=\s*(\[.*?\]);',
            ]:
                matches = re.findall(pattern, raw, re.DOTALL)
                if matches and matches[0].startswith("["):
                    try:
                        arr = json.loads(matches[0])
                        for item in arr:
                            if isinstance(item, dict):
                                issue = str(item.get("lotteryDrawNum", item.get("issue", "")))
                                date = str(item.get("lotteryDrawTime", item.get("date", "")))
                                nums_raw = item.get("lotteryDrawResult", "")
                                if isinstance(nums_raw, str):
                                    nums = [int(x) for x in nums_raw.split() if x.strip().isdigit()]
                                elif isinstance(nums_raw, list):
                                    nums = [int(x) for x in nums_raw]
                                else:
                                    nums = []
                                if issue and date and len(nums) >= 7:
                                    draws.append({
                                        "issue": issue,
                                        "date": date,
                                        "front": sorted(nums[:5]),
                                        "back": sorted(nums[5:7]),
                                    })
                    except json.JSONDecodeError:
                        pass

            if draws:
                draws.sort(key=lambda d: int(d["issue"]), reverse=True)
                print(f"    解析到 {len(draws)} 期")
                return draws

        except Exception as e:
            print(f"    失败: {type(e).__name__}")
            continue

    return draws


# ============ 核心获取逻辑 ============

def fetch_all():
    print(f"获取 {PERIOD} 期数据")
    print(f"使用 {'requests' if HAS_REQUESTS else 'urllib'} 库\n")

    all_draws = {}

    # 策略 1: API (首选, 连续无间断)
    print("--- 数据源 1: 中国体彩 API ---")
    api_draws = fetch_from_api(timeout=60)
    for d in api_draws:
        all_draws[d["issue"]] = d

    # 策略 2: 500.com (API 失败时使用)
    if len(all_draws) < PERIOD:
        print(f"\n--- 数据源 2: 500彩票网 ---")
        d500 = fetch_from_500(timeout=60)
        for d in d500:
            if d["issue"] not in all_draws:
                all_draws[d["issue"]] = d
        print(f"  合并后: {len(all_draws)} 期")

    # 策略 3: 官网页面 (兜底)
    if len(all_draws) < PERIOD:
        print(f"\n--- 数据源 3: 官网页面 ---")
        web_draws = fetch_from_webpage(timeout=60)
        for d in web_draws:
            if d["issue"] not in all_draws:
                all_draws[d["issue"]] = d
        print(f"  合并后: {len(all_draws)} 期")

    draws = sorted(all_draws.values(), key=lambda d: int(d["issue"]), reverse=True)
    return draws[:PERIOD]


# ============ 数据验证 ============

def verify_continuity(draws):
    """验证数据连续性"""
    if not draws:
        return False
    issues = [int(d["issue"]) for d in draws]
    diffs = [issues[i+1] - issues[i] for i in range(len(issues)-1)]
    gaps = [(issues[i], issues[i+1], d) for i, d in enumerate(diffs) if d > 1]
    if gaps:
        print(f"\n  ⚠ 数据存在 {len(gaps)} 个间断:")
        for b, a, d in gaps[:5]:
            print(f"    {b} -> {a} (缺 {d-1} 期)")
        if len(gaps) > 5:
            print(f"    ... 还有 {len(gaps)-5} 个")
        return False
    else:
        print(f"\n  ✓ 数据完全连续 ({len(draws)} 期)")
        return True


# ============ 保存逻辑 ============

def save_js(draws):
    path = os.path.join(OUT_DIR, "all_draws.js")
    lines = [
        f"// 大乐透开奖数据 — {len(draws)}期",
        f"// 自动更新: {time.strftime('%Y-%m-%d %H:%M:%S')} UTC",
        "// 数据来源: 中国体彩官方API / 500彩票网",
        "",
        "const ALL_DRAWS = [",
    ]
    for i, d in enumerate(reversed(draws)):
        front_str = " ".join(f"{n:02d}" for n in d["front"])
        back_str = " ".join(f"{n:02d}" for n in d["back"])
        line = (
            f'  {{issue:"{d["issue"]}",date:"{d["date"]}",'
            f'front:[{",".join(map(str, d["front"]))}],'
            f'back:[{",".join(map(str, d["back"]))}],'
            f'front_str:"{front_str}",back_str:"{back_str}"}}'
        )
        if i < len(draws) - 1:
            line += ","
        lines.append(line)
    lines.append("];\n")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    size = os.path.getsize(path)
    print(f"  all_draws.js: {size:,} bytes")
    return path


def save_xlsx(draws):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        path = os.path.join(OUT_DIR, "1.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "大乐透开奖数据"

        headers = [
            "期号", "开奖时间", "周",
            "前区一", "前区二", "前区三", "前区四", "前区五",
            "后区一", "后区二",
        ]

        hdr_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
        hdr_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        front_font = Font(name="微软雅黑", size=11, bold=True, color="C0392B")
        back_font = Font(name="微软雅黑", size=11, bold=True, color="2980B9")
        normal_font = Font(name="微软雅黑", size=10)
        center = Alignment(horizontal="center", vertical="center")
        thin = Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0"),
        )

        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center
            cell.border = thin

        weekdays = ["日", "一", "二", "三", "四", "五", "六"]
        for r, d in enumerate(draws, 2):
            ws.cell(row=r, column=1, value=d["issue"]).font = normal_font
            ws.cell(row=r, column=2, value=d["date"]).font = normal_font
            try:
                t = time.strptime(d["date"], "%Y-%m-%d")
                ws.cell(row=r, column=3, value=weekdays[t.tm_wday]).font = normal_font
            except Exception:
                pass
            for i, n in enumerate(d["front"]):
                ws.cell(row=r, column=4 + i, value=n).font = front_font
            for i, n in enumerate(d["back"]):
                ws.cell(row=r, column=9 + i, value=n).font = back_font
            for c in range(1, 11):
                ws.cell(row=r, column=c).alignment = center
                ws.cell(row=r, column=c).border = thin

        widths = [10, 12, 6, 6, 6, 6, 6, 6, 6, 6]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(draws) + 1}"

        wb.save(path)
        print(f"  1.xlsx: {os.path.getsize(path):,} bytes")
        return path
    except ImportError:
        print("  openpyxl 未安装，跳过 xlsx 生成")
        return None


def load_existing():
    """加载已有的 all_draws.js 数据"""
    path = os.path.join(OUT_DIR, "all_draws.js")
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            content = f.read()
        pattern = r'\{issue:"(\d+)",date:"([^"]+)",front:\[([^\]]+)\],back:\[([^\]]+)\]'
        existing = {}
        for m in re.finditer(pattern, content):
            issue = m.group(1)
            date = m.group(2)
            front = [int(x.strip()) for x in m.group(3).split(",") if x.strip()]
            back = [int(x.strip()) for x in m.group(4).split(",") if x.strip()]
            if len(front) == 5 and len(back) == 2:
                existing[issue] = {
                    "issue": issue,
                    "date": date,
                    "front": sorted(front),
                    "back": sorted(back),
                }
        print(f"  加载已有数据: {len(existing)} 期")
        return existing
    except Exception as e:
        print(f"  加载已有数据失败: {e}")
        return {}


if __name__ == "__main__":
    print("=" * 50)
    print("  大乐透数据自动更新")
    print("=" * 50)

    if not HAS_REQUESTS:
        print("  警告: 未安装 requests 库")

    try:
        existing = load_existing()
        draws = fetch_all()

        # 与现有数据合并 (现有数据只补充缺失的)
        if existing and len(draws) < PERIOD:
            for issue, d in existing.items():
                if issue not in {x["issue"] for x in draws}:
                    draws.append(d)
            draws.sort(key=lambda d: int(d["issue"]), reverse=True)
            draws = draws[:PERIOD]
            print(f"\n与现有数据合并后: {len(draws)} 期")

        print(f"\n共获取 {len(draws)} 期")

        # 验证数据连续性
        verify_continuity(draws)

        if draws:
            save_js(draws)
            save_xlsx(draws)
            print("\n完成！")
        else:
            print("\n警告: 未能获取新数据，保留现有数据")
    except Exception as e:
        import traceback
        print(f"\n获取数据失败: {e}")
        traceback.print_exc()
        print("保留现有数据")
        sys.exit(0)